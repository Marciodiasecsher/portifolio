import openpyxl
#Leitura e maniipulação dos dados do excel.
from selenium import webdriver
#Abrir o Navegador.
from selenium.webdriver.common.by import By
#Encontra elementos dentro da página e clica no botão.
from time import sleep
#Permite pausar.


variavel1 = openpyxl.load_workbook('planilha que irá tirar as informações.xlsx')
#abrir a planilha.
variavel2 = variavel1['Sheet#']
#Selecionando a página específica.
driver = webdriver.Chrome()
#Abre o Navegador.
driver.get('site')
#Acessa o site desejado.
for linha in variavel2.iter_rows(min_row=2,values_only=True):
#Ler cada linha da planilha, especificando qual linha iniciar a leitura.
    xxx, xxx, xxx, xxx = linha
    #percorre a linha com as informações específicas
   
    sleep(5)
     #Pausa por 5 segundos.
    variavel3 = driver.find_element(By.XPATH,"//[@='']")
    #Pesquisa o elemento epecífico da página
    #XPATH importante para a seleção do elemento correto na página
    sleep(1)
    #Pausa importante para o funcionamento correto
    variavel3.clear()
    #Apaga os caracters do campo de pesquisa.
    variavel3.send_keys(xxx)
    #função do Selenium para escrever.
    sleep(1)
    #Pausa importante para o funcionamento correto
    variaivel4 = driver.find_element(By.XPATH,"//[@='']")
    sleep(1)
    variavel4.click()
    #clica no elemento específico da página
    sleep(4)
    variavel5 = driver.find_element(By.XPATH,"//[@='']")
    if variavel5.text == 'nome do que você procura no campo do elemento':
        variiavel6 = driver.find_element(By.XPATH,"//[@='']")
        variavel7 = driver.find_element(By.XPATH,"//[@='']")
        variavel8 = variavel6.text.split()[3]
        #pega o item especifico na lista
        varaivel9 = variavel7.text.split()[3]
        variavel10 = openpyxl.load_workbook('planilha para listar informações retiradas do site.xlsx')
        variavel11 = variavel10['Sheet#']
        variavel11.append([xxx, xxx, xxx, xxx, 'nome do que você procura no elemento',variavel8, variavel9])
        #inclui a lista informações retirada do elemento
        variavel10.save('planilha que irá salvar informações retiradas do site.xlsx')
        #salva na planilha específica
    else:
        variavel10 = openpyxl.load_workbook('planilha para listar informações retiradas do site.xlsx')
        variavel11 = variavel10['Sheet#']
        variavel11.append([xxx, xxx, xxx, xxx,'informção retirada do site'])
        variavel10.save('planilha para listar informações retiradas do site.xlsx')
               #automatização webscrag